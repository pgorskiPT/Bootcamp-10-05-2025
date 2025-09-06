import glob
from pprint import pprint

from openpyxl.reader.excel import load_workbook
from openpyxl import load_workbook
# """Return a list of paths matching a pathname pattern.

file_list = glob.glob("*.xlsx")
print(file_list)
pprint(file_list)
# ['dane_zaktualizowane.xlsx',
#  'openpyxl_optimized.xlsx',
#  'tabela_przestawna.xlsx',
#  'videogamesales.xlsx',
#  'vgsales_formated.xlsx']

for file in file_list:
    try:
        wb = load_workbook(file)
        ws = wb.active

        value = ws['A1'].value
        print(f"{file}: A1 = {value}")
    except Exception as e:
        print("Bład:", e)

# dane_zaktualizowane.xlsx: A1 = Rank
# openpyxl_optimized.xlsx: A1 = 0
# tabela_przestawna.xlsx: A1 = Data
# videogamesales.xlsx: A1 = Rank
# vgsales_formated.xlsx: A1 = Rank
